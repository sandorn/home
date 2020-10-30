#  !/usr/bin/env python
#  -*- coding: utf-8 -*-
'''
==============================================================
Descripttion : None
Develop      : VSCode
Author       : Even.Sand
Contact      : sandorn@163.com
Date         : 2020-09-12 13:47:46
FilePath     : /py学习/excel乘法表.py
LastEditTime : 2020-09-16 10:00:26
Github       : https://github.com/sandorn/home
# ==============================================================
'''
# 此脚本用于在Excel文件中，生成一个乘法表，启动时输入模板：python 1207.py 9(乘法表中的最大基数)
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from sys import argv

shuzhi = 19
wb = openpyxl.Workbook()

sheet = wb.active

geshi = Font(bold=True)  # 设定格式，字体加粗

for i in range(1, int(shuzhi) + 1):  # 循环提取数值
    sheet['A' + str(i + 1)] = i  # 对A列中不同行数的单元格进行赋值
    sheet['A' + str(i + 1)].font = geshi  # 更改单元格格式为字体加粗

    lie1 = get_column_letter(i + 1)  # 转换列数为字母
    sheet[lie1 + str(1)] = i  # 对第1行中不同列数的单元格进行赋值
    sheet[lie1 + str(1)].font = geshi  # 更改单元格格式为字体加粗

for a in range(2, sheet.max_row + 1):  # 循环提取工作表中行数
    for b in range(2, sheet.max_column + 1):  # 循环提取工作表中列数
        lie2 = get_column_letter(b)  # 转换列数为字母
        sheet[lie2 + str(a)] = '=' + lie2 + str(1) + '*' + 'A' + str(a)  # 对单元格内容赋值，值为计算函数

wb.save('abc.xlsx')
