# !/usr/bin/env python
# -*- coding: utf-8 -*-
'''
==============================================================
Description  :
Develop      : VSCode
Author       : Even.Sand
Contact      : sandorn@163.com
Date         : 2020-11-26 19:38:49
LastEditTime : 2022-12-03 18:18:08
FilePath     : /xjLib/xt_File.py
Github       : https://github.com/sandorn/home
==============================================================
'''
import xlrd
import xlwt
import os
import openpyxl


def read_xls_excel(url, index):
    '''
    读取xls格式文件
    参数:
        url:文件路径
        index:工作表序号(第几个工作表,传入参数从1开始数)
    返回:
        data:表格中的数据
    '''
    # 打开指定的工作簿
    workbook = xlrd.open_workbook(url)
    # 获取工作簿中的所有表格
    sheets = workbook.sheet_names()
    # 获取工作簿中所有表格中的的第 index 个表格
    worksheet = workbook.sheet_by_name(sheets[index - 1])
    # 定义列表存储表格数据
    data = []
    # 遍历每一行数据
    for i in range(0, worksheet.nrows):
        # 定义表格存储每一行数据
        da = []
        # 遍历每一列数据
        for j in range(0, worksheet.ncols):
            # 将行数据存储到da列表
            da.append(worksheet.cell_value(i, j))
        # 存储每一行数据
        data.append(da)
    # 返回数据
    return data


def write_xls_excel(url,sheet_name,two_dimensional_data):
  '''
    写入xls格式文件
    参数:
        url:文件路径
        sheet_name:表名
        two_dimensional_data:将要写入表格的数据(二维列表)
    '''
    # 创建工作簿对象
    workbook = xlwt.Workbook()
    # 创建工作表对象
    sheet = workbook.add_sheet(sheet_name)
    # 遍历每一行数据
    for i in range(0,len(two_dimensional_data)):
        # 遍历每一列数据
        for j in range(0,len(two_dimensional_data[i])):
            # 写入数据
            sheet.write(i,j,two_dimensional_data[i][j])
    # 保存
    workbook.save(url)
    print("写入成功")

def write_xls_excel_add(url, two_dimensional_data, index):
    '''
    追加写入xls格式文件
    参数:
        url:文件路径
        two_dimensional_data:将要写入表格的数据(二维列表)
        index:指定要追加的表的序号(第几个工作表,传入参数从1开始数)
    '''
    # 打开指定的工作簿
    workbook = xlrd.open_workbook(url)
    # 获取工作簿中的所有表格
    sheets = workbook.sheet_names()
    # 获取指定的表
    worksheet = workbook.sheet_by_name(sheets[index-1])
    # 获取表格中已存在的数据的行数
    rows_old = worksheet.nrows
    # 将xlrd对象拷贝转化为xlwt对象
    new_workbook = copy(workbook)
    # 获取转化后工作簿中的第index个表格
    new_worksheet = new_workbook.get_sheet(index-1)
    # 遍历每一行数据
    for i in range(0, len(two_dimensional_data)):
        # 遍历每一列数据
        for j in range(0, len(two_dimensional_data[i])):
            # 追加写入数据,注意是从i+rows_old行开始写入
            new_worksheet.write(i+rows_old, j, two_dimensional_data[i][j])
    # 保存工作簿
    new_workbook.save(url)
    print("追加写入成功")

def read_xlsx_excel(url, sheet_name):
    '''
    读取xlsx格式文件
    参数:
        url:文件路径
        sheet_name:表名
    返回:
        data:表格中的数据
    '''
    # 使用openpyxl加载指定路径的Excel文件并得到对应的workbook对象
    workbook = openpyxl.load_workbook(url)
    # 根据指定表名获取表格并得到对应的sheet对象
    sheet = workbook[sheet_name]
    # 定义列表存储表格数据
    data = []
    # 遍历表格的每一行
    for row in sheet.rows:
        # 定义表格存储每一行数据
        da = []
        # 从每一行中遍历每一个单元格
        for cell in row:
            # 将行数据存储到da列表
            da.append(cell.value)
        # 存储每一行数据
        data.append(da)
    # 返回数据
    return data

def write_xlsx_excel(url, sheet_name, two_dimensional_data):
    '''
    写入xlsx格式文件
    参数:
        url:文件路径
        sheet_name:表名
        two_dimensional_data:将要写入表格的数据(二维列表)
    '''
    # 创建工作簿对象
    workbook = openpyxl.Workbook()
    # 创建工作表对象
    sheet = workbook.active
    # 设置该工作表的名字
    sheet.title = sheet_name
    # 遍历表格的每一行
    for i in range(0, len(two_dimensional_data)):
        # 遍历表格的每一列
        for j in range(0, len(two_dimensional_data[i])):
            # 写入数据(注意openpyxl的行和列是从1开始的,和我们平时的认知是一样的)
            sheet.cell(row=i + 1, column=j + 1, value=str(two_dimensional_data[i][j]))
    # 保存到指定位置
    workbook.save(url)
    print("写入成功")

def write_xlsx_excel_add(url, sheet_name, two_dimensional_data):
    '''
    追加写入xlsx格式文件
    参数:
        url:文件路径
        sheet_name:表名
        two_dimensional_data:将要写入表格的数据(二维列表)
    '''
    # 使用openpyxl加载指定路径的Excel文件并得到对应的workbook对象
    workbook = openpyxl.load_workbook(url)
    # 根据指定表名获取表格并得到对应的sheet对象
    sheet = workbook[sheet_name]
    for tdd in two_dimensional_data:
        sheet.append(tdd)
    # 保存到指定位置
    workbook.save(url)
    print("追加写入成功")
