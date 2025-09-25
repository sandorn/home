# !/usr/bin/env python3
"""
==============================================================
Description  : Excel操作工具 - 提供基于openpyxl的Excel文件读写功能
Develop      : VSCode
Author       : sandorn sandorn@live.cn
Date         : 2022-12-22 17:35:56
LastEditTime : 2024-09-06 11:00:00
FilePath     : /CODE/xjlib/xt_database/opxl.py
Github       : https://github.com/sandorn/home

本模块提供以下核心功能:
- ExcelHandler类:封装了openpyxl库的主要功能,简化Excel文件操作
- 支持Excel文件的创建、读取、修改和保存
- 支持工作表的创建、删除和切换
- 支持单元格、行、列和整个工作表数据的读写
- 支持数据以字典形式返回,方便处理结构化数据

主要特性:
- 支持上下文管理器,自动处理资源关闭
- 提供灵活的API,支持多种数据读写方式
- 包含错误处理和类型检查,提高代码健壮性
- 集成日志系统,便于调试和问题追踪
==============================================================
"""
from __future__ import annotations

import os
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from xt_wraps.log import LogCls

# 初始化日志实例
logger = LogCls()


def rename_file(file_path: str) -> str:
    """为文件重命名,添加'-opxl'后缀

    Args:
        file_path: 原始文件路径

    Returns:
        str: 添加后缀后的文件路径
    """
    file_dir, file_name = os.path.split(file_path)
    base_name, extension = os.path.splitext(file_name)
    return os.path.join(file_dir, f'{base_name}-opxl{extension}')


class ExcelHandler:
    """Excel文件操作类 - 提供对Excel xlsx文件的创建、读取、修改和保存功能

    该类封装了openpyxl库的主要功能,提供简洁的API接口操作Excel文件
    支持上下文管理器模式,确保资源正确释放

    Args:
        file: Excel文件路径
        sheet_name: 工作表名称,默认为None(使用活动工作表)

    Attributes:
        file: Excel文件路径
        wb: Workbook对象
        sh: 当前工作表对象
        sh_name_list: 所有工作表名称列表
        headers: 表头数据(通过read_header方法设置)

    Example:
        >>> # 创建实例并写入数据
        >>> with ExcelHandler('example.xlsx', 'Sheet1') as excel:
        ...     excel.write_cell(1, 1, 'Hello')
        ...     excel.write_cell(1, 2, 'World')
        >>>
        >>> # 读取数据
        >>> with ExcelHandler('example.xlsx') as excel:
        ...     headers = excel.read_header()
        ...     all_data = excel.read_all_dict()
    """
    def __init__(self, file: str, sheet_name: str | None = None):
        """初始化ExcelHandler实例

        Args:
            file: Excel文件路径
            sheet_name: 工作表名称,默认为None(使用活动工作表)
        """
        self.file = file
        logger.info(f'初始化ExcelHandler,文件路径: {self.file}')
        
        # 如果文件不存在,创建新文件
        if not os.path.exists(self.file):
            logger.info(f'文件不存在,创建新文件: {self.file}')
            wb = Workbook()
            wb.save(self.file)
            wb.close()

        # 加载工作簿
        self.wb = load_workbook(self.file)
        self.sh_name_list = self.wb.sheetnames

        # 设置当前工作表
        self.sh: Worksheet | None = None
        if sheet_name is None:
            self.sh = self.wb.active
        elif sheet_name not in self.sh_name_list:
            logger.info(f'工作表不存在,创建新工作表: {sheet_name}')
            self.sh = self.wb.create_sheet(sheet_name)
        else:
            self.sh = self.wb[sheet_name]
            logger.info(f'切换到工作表: {sheet_name}')
        
        if not isinstance(self.sh, Worksheet):
            logger.error('工作表类型错误')
            raise TypeError('工作表类型错误')
        # 初始化表头属性
        self.headers: tuple[Any, ...] | None = None

    def __enter__(self) -> ExcelHandler:
        """支持上下文管理器协议

        Returns:
            ExcelHandler: 当前实例
        """
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> bool:
        """支持上下文管理器协议,确保资源正确关闭

        Args:
            exc_type: 异常类型
            exc_val: 异常值
            exc_tb: 异常追踪

        Returns:
            bool: 如果没有异常发生,返回True
        """
        self.wb.close()
        logger.info(f'关闭Excel文件: {self.file}')
        return not exc_type

    def create_sheet(self, index: int | None = None, title: str | None = None) -> None:
        """创建新工作表

        Args:
            index: 工作表位置索引
            title: 工作表名称
        """
        self.wb.create_sheet(index=index, title=title)
        self.save_workbook()
        logger.info(f'创建新工作表: 索引={index}, 标题={title}')

    def remove_sheet(self, sheet_name: str) -> None:
        """删除指定工作表

        Args:
            sheet_name: 要删除的工作表名称
        """
        if sheet_name in self.sh_name_list:
            sheet = self.wb[sheet_name]
            self.wb.remove(sheet)
            self.save_workbook()
            self.sh_name_list = self.wb.sheetnames
            logger.info(f'删除工作表: {sheet_name}')
        else:
            logger.warning(f'工作表不存在: {sheet_name}')

    def _switch_sheet(self, sheet_name: str) -> None:
        """切换到指定工作表

        Args:
            sheet_name: 工作表名称
        """
        if sheet_name in self.sh_name_list:
            now_sh = self.wb[sheet_name]
            if isinstance(now_sh, Worksheet):
                self.sh = now_sh
                logger.info(f'切换到工作表: {sheet_name}')
            else:
                logger.error(f'工作表类型错误: {sheet_name}')
        else:
            logger.warning(f'工作表不存在: {sheet_name}')

    def read_header(self, sheet_name: str | None = None) -> tuple[Any, ...]:
        """获取表头信息

        Args:
            sheet_name: 工作表名称,默认为None(使用当前工作表)

        Returns:
            Tuple: 表头数据元组
        """
        if sheet_name is not None:
            self._switch_sheet(sheet_name)
        
        # 读取第一行作为表头
        header_row = tuple(self.sh.iter_rows(max_row=1, values_only=True))
        if header_row:
            self.headers = header_row[0]
            logger.info(f'读取表头成功,包含{len(self.headers)}列')
            return self.headers
        
        logger.warning('未读取到表头数据')
        return tuple()

    def read_all(self, sheet_name: str | None = None) -> list[list[Any]]:
        """读取工作表中所有数据

        Args:
            sheet_name: 工作表名称,默认为None(使用当前工作表)

        Returns:
            List[List[Any]]: 二维列表形式的表格数据
        """
        if sheet_name is not None:
            self._switch_sheet(sheet_name)
            
        data = [[cell.value for cell in row] for row in self.sh.rows]
        logger.info(f'读取所有数据成功,包含{len(data)}行')
        return data

    def read_row(self, row: int | str, sheet_name: str | None = None) -> list[Any]:
        """读取指定行数据

        Args:
            row: 行索引(整数)或行标识(如'A')
            sheet_name: 工作表名称,默认为None(使用当前工作表)

        Returns:
            List[Any]: 行数据列表
        """
        if sheet_name is not None:
            self._switch_sheet(sheet_name)
            
        row_data = [cell.value for cell in self.sh[row]]
        logger.info(f'读取行{row}数据成功,包含{len(row_data)}列')
        return row_data

    def read_row_dict(self, row: int | str, sheet_name: str | None = None) -> dict[str, Any]:
        """以字典形式读取指定行数据(表头为键)

        Args:
            row: 行索引(整数)或行标识(如'A')
            sheet_name: 工作表名称,默认为None(使用当前工作表)

        Returns:
            Dict[str, Any]: 字典形式的行数据
        """
        row_data = self.read_row(row, sheet_name)
        titles = self.read_header(sheet_name)
        result = dict(zip(titles, row_data, strict=False))
        logger.info(f'读取行{row}字典数据成功,包含{len(result)}个字段')
        return result

    def read_col(self, col: int | str, sheet_name: str | None = None) -> list[Any]:
        """读取指定列数据

        Args:
            col: 列索引(整数)或列标识(如'A')
            sheet_name: 工作表名称,默认为None(使用当前工作表)

        Returns:
            List[Any]: 列数据列表
        """
        if sheet_name is not None:
            self._switch_sheet(sheet_name)
            
        # 处理列标识
        col_index = col if isinstance(col, int) else self.get_column_index(col)
        
        # 读取列数据
        col_data_tuple = tuple(self.sh.iter_cols(min_col=col_index, max_col=col_index, values_only=True))
        col_data = list(col_data_tuple[0]) if col_data_tuple else []
        logger.info(f'读取列{col}数据成功,包含{len(col_data)}行')
        return col_data

    def read_all_dict(self, sheet_name: str | None = None) -> list[dict[str, Any]]:
        """以字典列表形式读取所有数据(表头为键)

        Args:
            sheet_name: 工作表名称,默认为None(使用当前工作表)

        Returns:
            List[Dict[str, Any]]: 字典列表形式的所有数据
        """
        if sheet_name is not None:
            self._switch_sheet(sheet_name)
            
        data = []
        rows = self.read_all()[1:]  # 跳过表头
        titles = self.read_header()
        
        for row in rows:
            data.append(dict(zip(titles, row, strict=False)))
            
        logger.info(f'读取所有字典数据成功,包含{len(data)}行记录')
        return data

    def read_cell(self, row: int, column: int) -> Any:
        """读取指定单元格数据

        Args:
            row: 行索引
            column: 列索引

        Returns:
            Any: 单元格值
        """
        value = self.sh.cell(row, column).value
        logger.info(f'读取单元格({row},{column})数据: {value}')
        return value

    def write_cell(self, row: int, column: int, value: Any) -> None:
        """写入数据到指定单元格

        Args:
            row: 行索引
            column: 列索引
            value: 要写入的值
        """
        self.sh.cell(row, column).value = value
        self.save_workbook()
        logger.info(f'写入单元格({row},{column})数据: {value}')

    def append(self, data_list: list[list[Any]]) -> None:
        """追加写入数据到工作表

        Args:
            data_list: 二维列表形式的数据
        """
        for row_data in data_list:
            self.sh.append(row_data)
        self.save_workbook()
        logger.info(f'追加写入数据成功,包含{len(data_list)}行')

    def save_workbook(self, excel_path: str | None = None) -> None:
        """保存工作簿

        Args:
            excel_path: 保存路径,默认为None(保存到原文件)
        """
        save_path = excel_path or self.file
        self.wb.save(save_path)
        logger.info(f'保存工作簿成功: {save_path}')

    @staticmethod
    def get_column_index(col_str: str) -> int:
        """将列标识转换为列索引

        Args:
            col_str: 列标识(如'A', 'B', 'AA')

        Returns:
            int: 列索引
        """
        return column_index_from_string(col_str)

    @staticmethod
    def get_column_letter(col_num: int) -> str:
        """将列索引转换为列标识

        Args:
            col_num: 列索引

        Returns:
            str: 列标识
        """
        return get_column_letter(col_num)


if __name__ == '__main__':
    # 示例代码,演示ExcelHandler的基本使用
    with ExcelHandler('d:/1.xlsx', 'Sheet1') as excel_handler:
        excel_handler.write_cell(1, 2, '2019年')
        
        # 读取单元格数据
        cell_value = excel_handler.read_cell(1, 1)
        logger.info(f'单元格(1,1)的值: {cell_value}')
        
        # 读取工作表列表
        logger.info(f'工作表列表: {excel_handler.sh_name_list}')
        
        # 读取列数据
        col_data = excel_handler.read_col(2)
        logger.info(f'第2列数据: {col_data}')
        
        # 追加数据
        excel_handler.append([['AAA', 'BBB', 'CCC', 'DDD', 'EEEE']])
        
        # 写入特定单元格
        excel_handler.write_cell(28, 1, 9876)
        
        # 读取所有数据
        all_data = excel_handler.read_all()
        logger.info(f'所有数据行数: {len(all_data)}')
        
        # 读取表头
        headers = excel_handler.read_header()
        logger.info(f'表头: {headers}')
        
        # 读取行字典数据
        row_dict = excel_handler.read_row_dict(1)
        logger.info(f'第1行字典数据: {row_dict}')

"""
Openpyxl库对excel的常规处理详解_sweet tea111的博客-CSDN博客_load_work openpyxl
https://blog.csdn.net/weixin_44390373/article/details/118153901

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
wb = load_workbook("xxx.xlsx")
ws = wb.active
-------------------------------------------------------------------------
#Border 边框Side边线
#eg:将A1:D5范围的边框显示
thin = Side(border_style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for row in ws[A1:D5]:
    for cell in row:
        cell.border = border
--------------------------------------------------------------------------
#Alignment 单元格对齐方式
#horizontal:代表水平方向，left:左对齐;center:居中对齐;right:右对齐;distributed:分散对齐;centerContinuous:跨列居中;justify:两端对齐
#vertical
#vertical:代表垂直方向，center:居中;top:靠上;bottom:靠下;justify:两端对齐;distributed:分散对齐
#wrap_text或者wrapText:自动换行
#eg:将单元格A3设置为水平右对齐，垂直居中，自动换行
align = Alignment(horizontal='right', vertical='center', wrap_text=True)
ws['A3'].alignment = align
------------------------------------------------------------------------------
#Font 字体设置
#name:字体名字
#size:字体大小
#color:字体颜色
#bold:加粗 True or False
#italic:斜体 True or False
#strike:删除线 True or False
#vertAlign = None 垂直对齐
#纵向对齐方式有3种：baseline superscript subscript
#underline = 'none':下划线
#eg
font = Font(size=11, bold=True, name='微软雅黑', color="FF0000")
WS['A3'].font = font
------------------------------------------------------------------------------
#PatternFill 填充
#fill_type:填充类型，若没有特别指定类型，则后续的参数都无效
#soild:纯色填充、其余类型：'none'/'darkDown'/'darkGray'/'darkGrid'/'darkHorizontal'等
#start_color:前景色 end_color:背景色
#eg

```python
fill = PatternFill(patternType="solid",start_color="33CCFF")
ws['A3'].fill = fill
-------------------------------------------------------------------------------
"""
