# !/usr/bin/env python
# -*- coding: utf-8 -*-
'''
==============================================================
Description  :
Develop      : VSCode
Author       : Even.Sand
Contact      : sandorn@163.com
Date         : 2022-12-07 17:18:41
LastEditTime : 2022-12-15 14:45:41
FilePath     : /xjLib/xt_Excel.py
Github       : https://github.com/sandorn/home
==============================================================
https://blog.csdn.net/qq_44614026/article/details/113275741
https://www.cnblogs.com/xcc-/p/xcc_08.html
https://blog.csdn.net/weixin_45417815/article/details/122015992
'''

import os

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def rename(file_path):
    (filepath, tempfilename) = os.path.split(file_path)
    (filename, extension) = os.path.splitext(tempfilename)
    return os.path.join(filepath, 'OpenPyXl-' + filename + extension)


class ExcelHandler():
    ''' openpyxl操作Excel xlsx文件的类 '''

    def __init__(self, file, sheet_name=None):
        '''初始化函数'''
        self.file = rename(file)
        if not os.path.exists(file):
            # 新建一个新工作表
            _wb = openpyxl.Workbook()
            _wb.save(self.file)
            _wb.close()

        self.wb = openpyxl.load_workbook(file)
        self.sh_name_list = self.wb.sheetnames

        if sheet_name is None: _sh = self.wb.active
        elif sheet_name not in self.sh_name_list:
            _sh = self.wb.create_sheet(sheet_name)
        else:
            _sh = self.wb[sheet_name]
        assert isinstance(_sh, Worksheet)
        self.sh = _sh

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.wb.close()
        return False if exc_type else True

    def createsheet(self, index=None, title=None):
        self.wb.create_sheet(index=index, title=title)
        self.save_wb()

    def removeSheet(self, sheetName):
        self.wb.remove(sheetName)
        self.save_wb()

    def read_header(self, sheet_name=None):
        '''获取表头'''
        # 表头信息为当前表单的标题,限制最大行为1,values_only= True返回单元格的值
        # 如果不加tuple,返回的只是一个对象
        if sheet_name is not None:
            _sh = self.wb[sheet_name]
            assert isinstance(_sh, Worksheet)
            self.sh = _sh

        self.headers = tuple(self.sh.iter_rows(max_row=1, values_only=True))[0]
        return self.headers

    def read_all(self, sheet_name=None):
        '''
        读取xlsx格式文件中的数据
        返回:
            data:表格中的数据
        '''
        if sheet_name is not None:
            _sh = self.wb[sheet_name]
            assert isinstance(_sh, Worksheet)
            self.sh = _sh

        data = [[cell.value for cell in row] for row in self.sh.rows]
        return data

    def read_row(self, row, sheet_name=None):
        '''读取某一行数据'''
        if sheet_name is not None:
            _sh = self.wb[sheet_name]
            assert isinstance(_sh, Worksheet)
            self.sh = _sh

        return [cell.value for cell in self.sh[row]]

    def read_row_dict(self, row, sheet_name=None):
        '''
        获取某一行数据,且将表头中的内容与数据结合展示（以字典的形式）
        如：{'序号':1,'会员卡号': '680021685898','机场名称':'上海机场'}
        '''
        row_data = self.read_row(row, sheet_name)
        titles = self.read_header()
        data_dict = dict(zip(titles, row_data))
        return data_dict

    def read_col(self, col, sheet_name=None):
        '''读取某一列数据'''
        if sheet_name is not None:
            _sh = self.wb[sheet_name]
            assert isinstance(_sh, Worksheet)
            self.sh = _sh
        _r = tuple(self.sh.iter_cols(min_col=col, max_col=col, values_only=True))[0]
        return [*_r]

    def read_all_dict(self, sheet_name=None):
        '''
        获取所有数据,且将表头中的内容与数据结合展示（以字典的形式）如：
        [ {'序号':1,'会员卡号': '680021685898','机场名称':'上海机场'}, ]
        '''
        if sheet_name is not None:
            _sh = self.wb[sheet_name]
            assert isinstance(_sh, Worksheet)
            self.sh = _sh

        data = []
        rows = self.read_all()[1:]
        titles = self.read_header()
        for row in rows:
            data_dict = dict(zip(titles, row))
            data.append(data_dict)
        return data

    def read_cell(self, row, column):
        '''读取单元格数据'''
        return self.sh.cell(row, column).value

    def write_cell(self, row, column, value):
        self.sh.cell(row, column).value = value
        self.save_wb()

    def append(self, data_list):
        '''
        追加写入xlsx格式文件
        参数:data_list:将要写入表格的数据(二维列表)
        '''
        for tdd in data_list:
            self.sh.append(tdd)
        self.save_wb()

    def save_wb(self, excel_path=None):
        if excel_path is not None: self.wb.save(excel_path)
        self.wb.save(self.file)

    @staticmethod
    def get_num(_col_str):
        return column_index_from_string(_col_str)

    @staticmethod
    def get_letter(_col_num):
        return get_column_letter(_col_num)


if __name__ == "__main__":
    with ExcelHandler("d:/1.xlsx", "Sheet1") as he:
        he.write_cell(1, 2, '2019年')
        rows = he.read_all()
        he.append(rows)
        # print(he.read_all_dict())
        # print(he.read_row(1))
        print(he.read_cell(1, 1))
        print(he.sh_name_list)
        # he2 = ExcelHandler("d:/2.xlsx")
        # he2.append(rows)
        print(he.read_col(2))
        # he.createsheet()
        he.append([
            [
                1111,
                2222,
                3333,
                4444,
                5555,
            ],
        ])
        he.write_cell(28, 1, 9876)
####################################################
# he = ExcelHandler("d:/1.xlsx", "Sheet1")
# he.write_cell(1, 2, '2019年')
# rows = he.read_all()
# # he.append(rows)
# # print(he.read_dict())
# print(he.read_header())
# # print(he.read_cell(1, 3))
# print(he.sh_name_list)
# he2 = ExcelHandler("d:/2.xlsx", "Sheet99")
# he2.append(rows)
'''
Openpyxl库对excel的常规处理详解_sweet tea111的博客-CSDN博客_load_work openpyxl
https://blog.csdn.net/weixin_44390373/article/details/118153901

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
wb = load_workbook("xxx.xlsx")
ws = wb.active
-------------------------------------------------------------------------
#Border 边框Side边线
#eg:将A1:D5范围的边框显示
thin = Side(border_style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for row in ws[A1:D5]:
    for cell in row:
        cell.border = border
--------------------------------------------------------------------------
#Alignment 单元格对齐方式
#horizontal:代表水平方向，left:左对齐;center:居中对齐;right:右对齐;distributed:分散对齐;centerContinuous:跨列居中;justify:两端对齐
#vertical
#vertical:代表垂直方向，center:居中;top:靠上;bottom:靠下;justify:两端对齐;distributed:分散对齐
#wrap_text或者wrapText:自动换行
#eg:将单元格A3设置为水平右对齐，垂直居中，自动换行
align = Alignment(horizontal='right', vertical='center', wrap_text=True)
ws['A3'].alignment = align
------------------------------------------------------------------------------
#Font 字体设置
#name:字体名字
#size:字体大小
#color:字体颜色
#bold:加粗 True or False
#italic:斜体 True or False
#strike:删除线 True or False
#vertAlign = None 垂直对齐
#纵向对齐方式有3种：baseline superscript subscript
#underline = 'none':下划线
#eg
font = Font(size=11, bold=True, name='微软雅黑', color="FF0000")
WS['A3'].font = font
------------------------------------------------------------------------------
#PatternFill 填充
#fill_type:填充类型，若没有特别指定类型，则后续的参数都无效
#soild:纯色填充、其余类型：'none'/'darkDown'/'darkGray'/'darkGrid'/'darkHorizontal'等
#start_color:前景色 end_color:背景色
#eg

```python
fill = PatternFill(patternType="solid",start_color="33CCFF")
ws['A3'].fill = fill
-------------------------------------------------------------------------------



'''
